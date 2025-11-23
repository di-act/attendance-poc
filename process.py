import argparse, traceback
from pathlib import Path
import json, re, os, sys
from datetime import datetime, timezone
from docx import Document
from typing import List, Dict, Any, Tuple, Optional
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Configuration
AWS_ACCESS_KEY_ID="ASIAYFXTYGAZFO6FPKGD"
AWS_SECRET_ACCESS_KEY="tzCigiitWKVIz/9juiZ7qsUKwYCt3vmM/a4+2l3N"
AWS_SESSION_TOKEN="IQoJb3JpZ2luX2VjEH4aDmFwLXNvdXRoZWFzdC0yIkcwRQIgayS99NreF/NsYd3jpx/d6ofm3fD128qyIFx8OahBtecCIQCshAnCpFT+o6I1MPaTDzy9VRijr8kxDL2m6SPxZd6wHCr5AghHEAAaDDU2MjA3ODE2NzA5MCIM1VQD98F3EeA07JQ6KtYC0GIaxoPbVxUsxgpi5OSHdOGgecAGxtGkjfL1W8VQEFrlR1RyNoNdWTwXUky+x0secMk8DSFIUG+duYVXWt/mnHV3UWmE2U7LQUtLScdafkeXPOnKxRxO8lTdsJaHjrTtrJYmjgPy4UNjitW+aS5UR3p5bs+0C9PeWTa+AP8YN/Yx/jYqjCN9q1xHFfNJCYjGcBgX4qqfHMvvsX2yrUYbsJh7uQqaR+XDaQoW0bwJUHgWZ4p+/Y9flYwmYsMF/XlD88UV5mhD39tZllIJ0XE/eWoxgXJi/sevrhL7JSVrPyMwkW5KQ3HwcZG+YxzoWZe3l8jb9tTzTgZthsMMJHbtuMRCjXqVM1gt61aCi5JSR6YR2qjzKArXS9+eSv1Q2P1ZvElphsj0vdhi3pXJf/LX2YkxHePxStKqdZaH2+UGcy8l57OB1SWIcIIeYty/DZUu+YGd/+IaMO2NjskGOqUBr3A0BfPwYctphHl+GM4d2iBJ1JGfzaSMJR3Pe7HgSTzOcCh7lPHHIDOzYq+tvRH9ZgKKuxKZIXSHfYOgXlSSi6TBcheibNwtz2SQSXziksY/v7Xqtrsrq3t0TJYzBz00sKiBU/eRfVjMgyVGx2J7ThIRtztdyWurz53GBCHYBoVBdr8t/B4DazuDezg4qlqkGknuSEvSVFvVLYDdjGhrgolhsBqt"
S3_BUCKET = "bhp-poc-bucket"
S3_BUCKET = "bhp-poc-bucket"
AWS_REGION = "ap-southeast-2" 
S3_ARN="arn:aws:s3:::bhp-poc-bucket"
S3_URI="arn:aws:s3:ap-southeast-2:562078167090:accesspoint/bhp-results"

class S3Uploader:
    """Upload files to AWS S3 bucket."""
    
    def __init__(self, bucket_name: str):
        import boto3
        self.s3_client = boto3.client(
            "s3",
            aws_access_key_id=AWS_ACCESS_KEY_ID,
            aws_secret_access_key=AWS_SECRET_ACCESS_KEY,
            aws_session_token=AWS_SESSION_TOKEN,
            region_name=AWS_REGION
        )
        self.bucket_name = bucket_name
    
    def upload_file(self, file_path: str, s3_key: str) -> None:
        """Upload a file to S3."""
        try:
            self.s3_client.upload_file(file_path, self.bucket_name, s3_key)
            print(f"✓ Uploaded {file_path} to s3://{self.bucket_name}/{s3_key}")
        except Exception as e:
            print(f"Error uploading {file_path} to S3: {e}")
            raise

class DataFrameMergeWithVariance:
    """Merge DataFrames with variance calculation and XLSX export."""
    
    def __init__(self, df1: pd.DataFrame, df1_name: str,
                 df2: pd.DataFrame, df2_name: str,
                 key_columns: list,
                 variance_threshold: float = 12.0):
        """
        Initialise merger.
        
        Args:
            df1: First DataFrame (e.g., agreement hours as reference)
            df1_name: Display name for df1 (e.g. SES-Invoice Data)
            df2: Second DataFrame (e.g., actual hours from attendance)
            df2_name: Display name for df2 (e.g. Attendance Summary Data)
            key_columns: List of column names to merge on
        """
        self.df1 = df1.copy()
        self.df2 = df2.copy()
        self.df1_name = df1_name
        self.df2_name = df2_name
        self.key_columns = key_columns
        self.merged_df = None
        self.variance_df = None
        self.variance_threshold = variance_threshold
    
    def merge_dataframes(self) -> pd.DataFrame:
        """ Merge two DataFrames on key columns."""
        # if ' ' not in self.key_columns:
        #     self.merged_df = pd.merge(self.df1, self.df2, 
        #                         on=self.key_columns, 
        #                         how='outer',
        #                         suffixes=(f'_{self.df1_name}', f'_{self.df2_name}'))
        # else:
        #     merge_keys = self.key_columns.split(" ")
        #     merge_key_list = [key.strip() for key in merge_keys]
        #     self.merged_df = pd.merge(self.df2, self.df1, 
        #                     on=merge_key_list, 
        #                     how='outer',
        #                     suffixes=(f'_{self.df1_name}', f'_{self.df2_name}'))
        if self.key_columns is None or len(self.key_columns)==0:
            raise ValueError(f"No key columns specified for merging. cannot map records.: {self.df2_name} & {self.df1_name}")
        else:
            self.merged_df = pd.merge(self.df1, self.df2, 
                            on=self.key_columns, 
                            how='outer',
                            suffixes=(f'_{self.df1_name}', f'_{self.df2_name}'))
        print(f"✓ Merged {len(self.df2)} + {len(self.df1)} records → {len(self.merged_df)} records")
        return self.merged_df
    
    def calculate_variance(self, 
                          df1_hours_col: str,
                          df2_hours_col: str,
                          variance_col: str = "variance_hours",
                          pct_col: str = "variance_pct") -> pd.DataFrame:
        """
        Calculate variance between two numeric columns.
        
        Args:
            df1_hours_col: Column from df1 with hours
            df2_hours_col: Column from df2 with hours
            variance_col: Output column for absolute variance
            pct_col: Output column for variance percentage
        """
        if self.merged_df is None:
            self.merge_dataframes()
        
        # Convert to numeric Prepare data for calculation
        self.merged_df[df1_hours_col] = pd.to_numeric(
            self.merged_df[df1_hours_col], errors='coerce').fillna(0)
        self.merged_df[df2_hours_col] = pd.to_numeric(
            self.merged_df[df2_hours_col], errors='coerce').fillna(0)
        
        # Calculate variance
        self.merged_df[variance_col] = (
            self.merged_df[df2_hours_col] - self.merged_df[df1_hours_col])
        
        # Calculate absolute variance
        self.merged_df[f'abs_{variance_col}'] = np.abs(self.merged_df[variance_col])
        
        # Calculate percentage (handle division by zero)
        actual = self.merged_df[df1_hours_col].to_numpy(dtype=float)
        allowed = self.merged_df[df2_hours_col].to_numpy(dtype=float)
        variance = self.merged_df[variance_col].to_numpy(dtype=float)
        
        with np.errstate(divide='ignore', invalid='ignore'):
            pct = np.where(
                allowed == 0,
                np.where(variance == 0, 0.0, np.nan),
                (variance / actual) * 100.0
            )
        
        self.merged_df[pct_col] = np.abs(np.round(pct, 2))
        
        print(f"✓ Calculated variance and percentages")
        return self.merged_df
    
    def get_variance_summary(self, variance_col: str = "variance_hours",
                            pct_col: str = "variance_pct") -> pd.DataFrame:
        """Generate summary DataFrame with key columns and variance."""
        if self.merged_df is None:
            return None
        # Prepare summary DataFrame with key columns and variance info
        summary_cols = self.key_columns + ['attendanceDate'] + ['totalSystemHours'] + ['totalHoursWorked'] + [
            variance_col, f'abs_{variance_col}', pct_col
        ]
        available_cols = [c for c in summary_cols if c in self.merged_df.columns]
        
        self.variance_df = self.merged_df[available_cols].copy()
        # self.variance_df = self.variance_df.rename(columns={
        #     'hoursWorked': 'systemHours'  # assuming hoursWorked from df1 is allowed
        # })
        self.variance_df = self.variance_df[self.variance_df['attendanceDate'].notna()]
        self.variance_df["hours_mismatch"] = np.abs(self.variance_df["variance_pct"]) > 10.0
        self.variance_df["policy_error"] = self.variance_df["totalHoursWorked"] > self.variance_threshold
        return self.variance_df
    
    def export_to_csv(self, output_file: str) -> None:
        if self.merged_df is None:
            self.merge_dataframes()
    
    def export_to_xlsx(self, output_file: str,
                      df1_hours_col: str = None,
                      df2_hours_col: str = None) -> None:
        """
        Export all three DataFrames to XLSX with merged headers.
        
        Args:
            output_file: Output XLSX file path
            df1_hours_col: Hours column from df1
            df2_hours_col: Hours column from df2
        """
        if self.merged_df is None:
            self.merge_dataframes()
        
        if self.variance_df is None and df1_hours_col and df2_hours_col:
            self.calculate_variance(df1_hours_col, df2_hours_col)
            self.get_variance_summary()
        
        # Create Excel file with multiple sheets
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            # Sheet 1: Original df1
            self.df1.to_excel(writer, sheet_name=self.df1_name, 
                            startrow=2, index=False)
            
            # Sheet 2: Original df2
            self.df2.to_excel(writer, sheet_name=self.df2_name, 
                            startrow=2, index=False)
            
            # Sheet 3: Merged data
            self.merged_df.to_excel(writer, sheet_name='Merged', 
                                   startrow=2, index=False)
            
            # Sheet 4: Variance summary
            if self.variance_df is not None:
                self.variance_df.to_excel(writer, sheet_name='Comparison', 
                                         startrow=2, index=False)

        # Format all sheets
        wb = openpyxl.load_workbook(output_file)
        
        self._format_sheet(wb[self.df1_name], self.df1_name, len(self.df1.columns))
        wb[self.df1_name].sheet_properties.tabColor = 'F2AA84' # Dark Brown
        self._format_sheet(wb[self.df2_name], self.df2_name, len(self.df2.columns))
        wb[self.df2_name].sheet_properties.tabColor = '538DD5' # Dark Blue
        self._format_sheet(wb['Merged'], 'Merged Data', len(self.merged_df.columns))
        wb['Merged'].sheet_properties.tabColor = 'CCC0DA' # Dark Blue
        
        
        if self.variance_df is not None:
            self._format_sheet(wb['Comparison'], 'Variance Analysis', 
                             len(self.variance_df.columns))
        # format Comparison sheet
        comp_worksheet = wb['Comparison']
        fill = PatternFill(start_color="538DD5", end_color="538DD5", 
                              fill_type="solid")
        cell = comp_worksheet.cell(row=2, column=5)
        cell.fill = fill
        # delete variance_hours column
        variance_col_idx = 6
        comp_worksheet.delete_cols(variance_col_idx)
        purple_fill = PatternFill(start_color="CCC0DA", end_color="CCC0DA", 
                              fill_type="solid")
        for col in range(6, 10):
            cell = comp_worksheet.cell(row=2, column=col)
            cell.fill = purple_fill
        comp_worksheet.merge_cells('A1:I1')
        comp_worksheet.delete_cols(10)
        
        # renaming columns
        comp_worksheet.cell(row=3, column=1).value = "UID"
        comp_worksheet.cell(row=3, column=2).value = "Services Performed"
        comp_worksheet.cell(row=3, column=3).value = "Service Date"
        comp_worksheet.cell(row=3, column=4).value = "Invoice Hours"
        comp_worksheet.cell(row=3, column=5).value = "System Hours"
        comp_worksheet.cell(row=3, column=6).value = "Variance (Hours)"
        comp_worksheet.cell(row=3, column=7).value = "Variance (Percentage)"
        comp_worksheet.cell(row=3, column=8).value = "Mismatch Hours"
        comp_worksheet.cell(row=3, column=9).value = "Policy Conflict"
        comp_worksheet.column_dimensions['B'].width = 50
        # Activate dashboard sheet
        wb.active = wb.sheetnames.index("Comparison")
        wb.save(output_file)
       
        print(f"✓ XLSX file created: {output_file}")
        print(f"  - Sheet1 : {self.df1_name} ({len(self.df1)} records)")
        print(f"  - Sheet2 : {self.df2_name} ({len(self.df2)} records)")
        print(f"  - Merged : Combined data ({len(self.merged_df)} records)")
        if self.variance_df is not None:
            print(f"  - Comparison : Summary ({len(self.variance_df)} records)")
    
    def _format_sheet(self, worksheet, header_name: str, num_cols: int) -> None:
        """Format worksheet with merged headers and styling."""
        # Merge cells for main header
        if header_name == 'Variance Analysis':
            worksheet.merge_cells(f'A1:I1')
        else:
            worksheet.merge_cells(f'A1:{get_column_letter(num_cols)}1')
        header_cell = worksheet['A1']
        header_cell.value = header_name
        
        # Header styling
        header_fill = PatternFill(start_color="D45A16", end_color="D45A16", 
                                 fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=12)
        center_align = Alignment(horizontal="center", vertical="center")
        
        header_cell.fill = header_fill
        header_cell.font = header_font
        header_cell.alignment = center_align
        worksheet.row_dimensions[1].height = 25
        
        # Column header styling (row 2)
        sub_fill = PatternFill(start_color="F2AA84", end_color="F2AA84", 
                              fill_type="solid")
        sub_font = Font(bold=True, size=11)
        
        for col in range(1, num_cols + 1):
            cell = worksheet.cell(row=2, column=col)
            cell.fill = sub_fill
            cell.font = sub_font
            cell.alignment = center_align
        
        # Auto-adjust column widths
        for col in range(1, num_cols + 1):
            worksheet.column_dimensions[get_column_letter(col)].width = 20
        
        # Add borders
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row,
                                       min_col=1, max_col=num_cols):
            for cell in row:
                cell.border = thin_border

# TASK-1:: Extract data from DOCX agreement file
def extract_agreement_data(docx_filepath: str) -> Dict[str, Any]:
    """
    Extract structured data from a DOCX agreement file and return as JSON.
    
        UID: 88888888
        system hours: 7
        services performed: electrical, engineering, plumbing

    Args:
        docx_filepath: Path to the DOCX file
        
    Returns:
        Dictionary with extracted data in the required format
    """
     # Check file exists
    if not os.path.exists(docx_filepath) or not os.path.isfile(docx_filepath):
        raise FileNotFoundError(f"DOCX file not found: {docx_filepath}")
    
    
    # Read the DOCX file
    doc = Document(docx_filepath)
    
    # Extract all text from the document
    full_text = "\n".join([para.text for para in doc.paragraphs])
    
    # Split records by UID pattern
    records = []
    uid_pattern = r"UID:\s*(\d+)"
    hours_pattern = r"system\s+hours:\s*(\d+)"
    services_pattern = r"services\s+performed:\s*(.*\n)"
    
    # Find all UID positions
    uid_matches = list(re.finditer(uid_pattern, full_text, re.IGNORECASE))
    
    for i, uid_match in enumerate(uid_matches):
        # Get the text section for this record
        start_pos = uid_match.start()
        end_pos = uid_matches[i + 1].start() if i + 1 < len(uid_matches) else len(full_text)
        record_text = full_text[start_pos:end_pos]
        
        # Extract fields for this record
        uid_match_obj = re.search(uid_pattern, record_text, re.IGNORECASE)
        hours_match_obj = re.search(hours_pattern, record_text, re.IGNORECASE)
        services_match_obj = re.search(services_pattern, record_text, re.IGNORECASE)
        
        if uid_match_obj:
            record = {
                "uid": int(uid_match_obj.group(1)),
                "systemHours": int(hours_match_obj.group(1)) if hours_match_obj else None,
                "servicesPerformed": services_match_obj.group(1) if services_match_obj else None
            }
            if record["systemHours"] is not None and record["servicesPerformed"] is not None:
                record["servicesPerformed"] = record["servicesPerformed"].strip() # clean up whitespace
                records.append(record)
    
    # Create the final JSON structure
    result = {
        "documentName": "agreement.docx",
        "totalRecord": len(records),
        "records": records
    }
    
    return result

# TASK-1.1::  Assume PatternExtractor is defined elsewhere and imported to process documents
def parse_document_text(text):
    """
    Contains the custom business logic to parse the document text
    and extract structured data using regular expressions.
    """
    
    # This regex looks for the three-line pattern,
    # capturing the data for each field.
    # It accounts for potential whitespace and newlines between blocks.
    pattern = re.compile(
        r"UID:\s*(\d+)\s*\n+"
        r"system hours:\s*(\d+)\s*\n+"
        r"services performed:\s*([\w\s]+)",
        re.IGNORECASE | re.MULTILINE
    )
    
    records=[] # List to hold all extracted records
    
    for match in pattern.finditer(text):
        try:
            record = {
                "uid": int(match.group(1)),
                "systemHours": int(match.group(2)),
                "servicesPerformed": match.group(3).strip()
            }
            records.append(record)
        except (ValueError, TypeError):
            # Log and skip malformed records
            print(f"Skipping malformed record: {match.groups()}")
            
    # Construct the final JSON object in the user-specified format
    output_json = {
        "documentName": "attendance",
        "totalRecord": len(records),
        "records": records
    }
    
    return output_json

# TASK-2:: Extract data from CSV attendance file
def extract_attendance_data(csv_filepath: str) -> Dict[str, Any]:
    """
    Extract structured data from a CSV attendance file and return as JSON.
    """
    # Check file exists
    if not os.path.exists(csv_filepath) or not os.path.isfile(csv_filepath):
        raise FileNotFoundError(f"CSV file not found: {csv_filepath}")

    # Read the CSV file to data frame
    df = pd.read_csv(csv_filepath)

    # Validate required columns
    required_columns = ["uid", "punchInDateTime", "punchOutDateTime", "servicesPerformed"]
    for col in required_columns:
        if col not in df.columns:
            raise ValueError(f"Missing required column in CSV: {col}")

    # Convert DataFrame to list of records
    records = df.to_dict(orient="records")

    # constructing document summary
    doc_summary = {
        "documentName": "attendance",
        "totalRecord": len(records),
        "records": records
    }

    # Ensure datetime columns are parsed correctly
    df["punchInDateTime"] = pd.to_datetime(df["punchInDateTime"].str.strip())
    df["punchOutDateTime"] = pd.to_datetime(df["punchOutDateTime"].str.strip())

    # Calculate hours worked
    df["hoursWorked"] = (df["punchOutDateTime"] - df["punchInDateTime"]).dt.total_seconds() / 3600

    # Extract attendance date
    df["attendanceDate"] = df["punchInDateTime"].dt.date

    # Group by UID and attendanceDate
    grouped = df.groupby(["uid", "attendanceDate"]).agg({
        "hoursWorked": "sum",
        "servicesPerformed": lambda x: ", ".join(x.str.strip())
    }).reset_index()

    # Rename columns
    grouped.rename(columns={
        "hoursWorked": "totalHoursWorked"
    }, inplace=True)

    return grouped

# region: Finally:: compare two data sets and generate summary with attendance records and flag the record to review


# define set of static variables read from command line and environment .env file
DOC_FILE = os.getenv("DOC_FILE")
OUTPUT_FILE = os.getenv("OUTPUT_FILE")
INVOICE_FILE = os.getenv("INVOICE_FILE")
ATTENDANCE_FILE = os.getenv("ATTENDANCE_FILE")



# get current root directory
ROOT_DIR = Path(__file__).resolve().parent




# Logging
basedir = os.path.abspath(os.path.dirname(__file__))
log_dir = os.path.join(basedir, "logs")
os.makedirs(log_dir, exist_ok=True)

accesslog = os.path.join(log_dir, "access.log")
errorlog = os.path.join(log_dir, "error.log")
loglevel = "info"
# endregion

def main():
    # region:: process prompt (Parse command line arguments)
    parser = argparse.ArgumentParser(
        description='Extract patterns from DOCX files'
    )

    # define command line arguments
    parser.add_argument(
        '-p', '--patterns',
        nargs='+',
        choices=['numeric', 'currency', 'datetime', 'email', 'phone', 'url', 'string'],
        help='Pattern types to extract (default: all)'
    )
    parser.add_argument(
        '-i', '--invoice',
        help='Input file (DOCX format, containing uid, systemHours, servicesPerformed)'
    )
    parser.add_argument(
        '-a', '--attendance',
        help='Input file (CSV format, containing uid, punchInDateTime, punchOutDateTime, servicesPerformed)'
    )
    parser.add_argument(
        '-s', '--summary',
        help='Summary Attendance file (CSV format with uid, attendanceDate, totalHoursWorked, servicesPerformed)'
    )
    
    args = parser.parse_args()
    # endregion
    try:
        # region:: extract data (DOCX, CSV) from input files
        # defining empty lists to hold extracted data
        agreement_ref=[]
        daily_attendance=[]
        attendance_summary=[]
        
        # Extract DOCX invoice file
        file_type = os.path.splitext(args.invoice)[1].lower()
        if file_type != '.docx':
            raise ValueError("Only DOCX files are supported.")
        else:
            file_path = os.path.join(ROOT_DIR, "input", args.invoice)
            agreement_ref = extract_agreement_data(file_path) # TASK-1

        # Extract CSV attendance file
        file_type = os.path.splitext(args.attendance)[1].lower()
        if file_type != '.csv':
            raise ValueError("Only CSV files are supported.")
        else:
            file_path = os.path.join(ROOT_DIR, "input", args.attendance)
            daily_attendance_summary_df = extract_attendance_data(file_path) # TASK-2

        # Task-3:: Convert agreement records to DataFrame for comparison
        agreement_ref_df = pd.DataFrame(agreement_ref['records'])

        # Group agreement data by uid and servicesPerformed to get max allowed hours and remove duplicates
        agreement_grouped = agreement_ref_df.groupby(['uid', 'servicesPerformed']).agg(
            totalSystemHours = ('systemHours', 'sum'), # assuming systemHours in agreement is the max allowed
            allServicesPerformed = ('servicesPerformed', lambda x: ', '.join(x.unique()))
        ).reset_index()
        # agreement_grouped.rename(columns={
        #     "hoursWorked": "totalHoursWorked"
        # }, inplace=True)
        # endregion
        
        # # Optionally save to JSON file with time suffix to avoid overwriting
        timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
        file_prefix = os.path.join(ROOT_DIR, "output")
        if not os.path.exists(file_prefix):
            os.makedirs(file_prefix)
        # # Save the generated DataFrame to an Excel file
        # agreement_grouped.to_excel(f"{file_prefix}\\tmp_agreement_grouped.xlsx", index=False)
        
        # test:: 
        # daily_attendance_summary_df.to_excel(f"{file_prefix}\\tmp_daily_attendance_summary.xlsx", index=False)
        
        # region:: Generating Output: Merge daily attendance summary with agreement grouped data on uid and servicesPerformed 
        # Create merger and process (AS GENERIC CLASS DEFINED IN data_merge_with_variance.py NOTE: parameter, variables are hardcoded for now)
        merger = DataFrameMergeWithVariance(daily_attendance_summary_df, "Time & Attendance Summary", # df1
                                           agreement_grouped, "SES-Invoice",                          # df2
                                           ['uid', 'servicesPerformed'])                              # key columns
        merger.merge_dataframes()
        file = os.path.join(file_prefix, "tmp_merged.csv")
        # test::
        # merger.merged_df.to_csv(file)
        merger.calculate_variance("totalHoursWorked", "totalSystemHours")
        # test::
        merger.merged_df.to_csv(file)
        
        merger.get_variance_summary()
        # test::
        file = os.path.join(file_prefix, "tmp_comparison.csv")
        merger.variance_df.to_csv(file)
        file = os.path.join(file_prefix, "result.xlsx")
        merger.export_to_xlsx(file, "totalHoursWorked", "totalSystemHours")
        # endregion
        
        print("✓ Data extraction and processing completed successfully.")
        
        # Initialize uploader
        try:
            uploader = S3Uploader(S3_URI)
        except Exception as e:
            print(f"\n✗ Failed to initialize S3 client: {e}")
            return False
        
        # Upload file
        upload_file_path = file
        success = uploader.upload_file(
            file_path=upload_file_path,
            s3_key="result.xlsx"
        )
        
        if not success:
            return False
    
        print(f"✓ Uploaded to s3 bucket -> https://bhp-poc-bucket.s3.ap-southeast-2.amazonaws.com/result.xlsx successfully.")
        # output_file = f"{ROOT_DIR}\\output\\output_{timestamp}.json"
        # with open(output_file, "w") as f:
        #     json.dump(daily_attendance_summary, f, indent=4)

    except FileNotFoundError as e:
        print(f"Error: {e}\n\n", file=sys.stderr)
        #traceback.print_exc()
        sys.exit(1)
    except Exception as e:
        print(f"Error: {e}\n\n", file=sys.stderr)
        #traceback.print_exc()
        sys.exit(1)

ROOT_DIR = Path(__file__).resolve().parent
if __name__ == "__main__":
    main()
