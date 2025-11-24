
import numpy as np
import openpyxl
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import pandas as pd


class DataFrameMergeWithVariance:
    """Merge DataFrames with variance calculation and XLSX export."""
    
    def __init__(self, df1: pd.DataFrame, df1_name: str,
                 df2: pd.DataFrame, df2_name: str,
                 key_columns: list,
                 variance_threshold: float = 12.0):
        """
        Initialise merger.
        
        Args:
            df1: First DataFrame (e.g., agreement hours as reference)
            df1_name: Display name for df1 (e.g. SES-Invoice Data)
            df2: Second DataFrame (e.g., actual hours from attendance)
            df2_name: Display name for df2 (e.g. Attendance Summary Data)
            key_columns: List of column names to merge on
        """
        self.df1 = df1.copy()
        self.df2 = df2.copy()
        self.df1_name = df1_name
        self.df2_name = df2_name
        self.key_columns = key_columns
        self.merged_df = None
        self.variance_df = None
        self.variance_threshold = variance_threshold
    
    def merge_dataframes(self) -> pd.DataFrame:
        """ Merge two DataFrames on key columns."""
        # if ' ' not in self.key_columns:
        #     self.merged_df = pd.merge(self.df1, self.df2, 
        #                         on=self.key_columns, 
        #                         how='outer',
        #                         suffixes=(f'_{self.df1_name}', f'_{self.df2_name}'))
        # else:
        #     merge_keys = self.key_columns.split(" ")
        #     merge_key_list = [key.strip() for key in merge_keys]
        #     self.merged_df = pd.merge(self.df2, self.df1, 
        #                     on=merge_key_list, 
        #                     how='outer',
        #                     suffixes=(f'_{self.df1_name}', f'_{self.df2_name}'))
        if self.key_columns is None or len(self.key_columns)==0:
            raise ValueError(f"No key columns specified for merging. cannot map records.: {self.df2_name} & {self.df1_name}")
        else:
            self.merged_df = pd.merge(self.df1, self.df2, 
                            on=self.key_columns, 
                            how='outer',
                            suffixes=(f'_{self.df1_name}', f'_{self.df2_name}'))
        print(f"✓ Merged {len(self.df2)} + {len(self.df1)} records → {len(self.merged_df)} records")
        return self.merged_df
    
    def calculate_variance(self, 
                          df1_hours_col: str,
                          df2_hours_col: str,
                          variance_col: str = "variance_hours",
                          pct_col: str = "variance_pct") -> pd.DataFrame:
        """
        Calculate variance between two numeric columns.
        
        Args:
            df1_hours_col: Column from df1 with hours
            df2_hours_col: Column from df2 with hours
            variance_col: Output column for absolute variance
            pct_col: Output column for variance percentage
        """
        if self.merged_df is None:
            self.merge_dataframes()
        
        # Convert to numeric Prepare data for calculation
        self.merged_df[df1_hours_col] = pd.to_numeric(
            self.merged_df[df1_hours_col], errors='coerce').fillna(0)
        self.merged_df[df2_hours_col] = pd.to_numeric(
            self.merged_df[df2_hours_col], errors='coerce').fillna(0)
        
        # Calculate variance
        self.merged_df[variance_col] = (
            self.merged_df[df2_hours_col] - self.merged_df[df1_hours_col])
        
        # Calculate absolute variance
        self.merged_df[f'abs_{variance_col}'] = np.abs(self.merged_df[variance_col])
        
        # Calculate percentage (handle division by zero)
        actual = self.merged_df[df1_hours_col].to_numpy(dtype=float)
        allowed = self.merged_df[df2_hours_col].to_numpy(dtype=float)
        variance = self.merged_df[variance_col].to_numpy(dtype=float)
        
        with np.errstate(divide='ignore', invalid='ignore'):
            pct = np.where(
                allowed == 0,
                np.where(variance == 0, 0.0, np.nan),
                (variance / actual) * 100.0
            )
        
        self.merged_df[pct_col] = np.abs(np.round(pct, 2))
        
        print(f"✓ Calculated variance and percentages")
        return self.merged_df
    
    def get_variance_summary(self, variance_col: str = "variance_hours",
                            pct_col: str = "variance_pct") -> pd.DataFrame:
        """Generate summary DataFrame with key columns and variance."""
        if self.merged_df is None:
            return None
        # Prepare summary DataFrame with key columns and variance info
        summary_cols = self.key_columns + ['attendanceDate'] + ['totalSystemHours'] + ['totalHoursWorked'] + [
            variance_col, f'abs_{variance_col}', pct_col
        ]
        available_cols = [c for c in summary_cols if c in self.merged_df.columns]
        
        self.variance_df = self.merged_df[available_cols].copy()
        # self.variance_df = self.variance_df.rename(columns={
        #     'hoursWorked': 'systemHours'  # assuming hoursWorked from df1 is allowed
        # })
        self.variance_df = self.variance_df[self.variance_df['attendanceDate'].notna()]
        self.variance_df["hours_mismatch"] = np.abs(self.variance_df["variance_pct"]) > 10.0
        self.variance_df["policy_error"] = self.variance_df["totalHoursWorked"] > self.variance_threshold
        return self.variance_df
    
    def export_to_csv(self, output_file: str) -> None:
        if self.merged_df is None:
            self.merge_dataframes()
    
    def export_to_xlsx(self, output_file: str,
                      df1_hours_col: str = None,
                      df2_hours_col: str = None) -> None:
        """
        Export all three DataFrames to XLSX with merged headers.
        
        Args:
            output_file: Output XLSX file path
            df1_hours_col: Hours column from df1
            df2_hours_col: Hours column from df2
        """
        if self.merged_df is None:
            self.merge_dataframes()
        
        if self.variance_df is None and df1_hours_col and df2_hours_col:
            self.calculate_variance(df1_hours_col, df2_hours_col)
            self.get_variance_summary()
        
        # Create Excel file with multiple sheets
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            # Sheet 1: Original df1
            self.df1.to_excel(writer, sheet_name=self.df1_name, 
                            startrow=2, index=False)
            
            # Sheet 2: Original df2
            self.df2.to_excel(writer, sheet_name=self.df2_name, 
                            startrow=2, index=False)
            
            # Sheet 3: Merged data
            self.merged_df.to_excel(writer, sheet_name='Calculation', 
                                   startrow=2, index=False)
            
            # Sheet 4: Variance summary
            if self.variance_df is not None:
                self.variance_df.to_excel(writer, sheet_name='Comparison', 
                                         startrow=2, index=False)

        # Format all sheets
        wb = openpyxl.load_workbook(output_file)
        
        self._format_sheet(wb[self.df1_name], self.df1_name, len(self.df1.columns))
        wb[self.df1_name].sheet_properties.tabColor = 'F2AA84' # Dark Brown
        self._format_sheet(wb[self.df2_name], self.df2_name, len(self.df2.columns))
        wb[self.df2_name].sheet_properties.tabColor = '538DD5' # Dark Blue
        self._format_sheet(wb['Calculation'], 'Calculated Data', len(self.merged_df.columns))
        wb['Calculation'].sheet_properties.tabColor = 'CCC0DA' # Dark Blue
        
        
        if self.variance_df is not None:
            self._format_sheet(wb['Comparison'], 'Variance Analysis', 
                             len(self.variance_df.columns))
        # format Comparison sheet
        comp_worksheet = wb['Comparison']
        fill = PatternFill(start_color="538DD5", end_color="538DD5", 
                              fill_type="solid")
        cell = comp_worksheet.cell(row=2, column=5)
        cell.fill = fill
        # delete variance_hours column
        variance_col_idx = 6
        comp_worksheet.delete_cols(variance_col_idx)
        purple_fill = PatternFill(start_color="CCC0DA", end_color="CCC0DA", 
                              fill_type="solid")
        for col in range(6, 10):
            cell = comp_worksheet.cell(row=2, column=col)
            cell.fill = purple_fill
        comp_worksheet.merge_cells('A1:I1')
        comp_worksheet.delete_cols(10)
        
        # renaming columns
        comp_worksheet.cell(row=3, column=1).value = "UID"
        comp_worksheet.cell(row=3, column=2).value = "Services Performed"
        comp_worksheet.cell(row=3, column=3).value = "Service Date"
        comp_worksheet.cell(row=3, column=4).value = "Invoice Hours"
        comp_worksheet.cell(row=3, column=5).value = "System Hours"
        comp_worksheet.cell(row=3, column=6).value = "Variance (Hours)"
        comp_worksheet.cell(row=3, column=7).value = "Variance (Percentage)"
        comp_worksheet.cell(row=3, column=8).value = "Mismatch Hours"
        comp_worksheet.cell(row=3, column=9).value = "Policy Conflict"
        comp_worksheet.column_dimensions['B'].width = 50
        # Activate dashboard sheet
        wb.active = wb.sheetnames.index("Comparison")
        wb.save(output_file)
       
        print(f"✓ XLSX file created: {output_file}")
        print(f"  - Sheet1 : {self.df1_name} ({len(self.df1)} records)")
        print(f"  - Sheet2 : {self.df2_name} ({len(self.df2)} records)")
        print(f"  - Merged : Combined data ({len(self.merged_df)} records)")
        if self.variance_df is not None:
            print(f"  - Comparison : Summary ({len(self.variance_df)} records)")
    
    def _format_sheet(self, worksheet, header_name: str, num_cols: int) -> None:
        """Format worksheet with merged headers and styling."""
        # Merge cells for main header
        if header_name == 'Variance Analysis':
            worksheet.merge_cells(f'A1:I1')
        else:
            worksheet.merge_cells(f'A1:{get_column_letter(num_cols)}1')
        header_cell = worksheet['A1']
        header_cell.value = header_name
        
        # Header styling
        header_fill = PatternFill(start_color="D45A16", end_color="D45A16", 
                                 fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=12)
        center_align = Alignment(horizontal="center", vertical="center")
        
        header_cell.fill = header_fill
        header_cell.font = header_font
        header_cell.alignment = center_align
        worksheet.row_dimensions[1].height = 25
        
        # Column header styling (row 2)
        sub_fill = PatternFill(start_color="F2AA84", end_color="F2AA84", 
                              fill_type="solid")
        sub_font = Font(bold=True, size=11)
        
        for col in range(1, num_cols + 1):
            cell = worksheet.cell(row=2, column=col)
            cell.fill = sub_fill
            cell.font = sub_font
            cell.alignment = center_align
        
        # Auto-adjust column widths
        for col in range(1, num_cols + 1):
            worksheet.column_dimensions[get_column_letter(col)].width = 20
        
        # Add borders
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row,
                                       min_col=1, max_col=num_cols):
            for cell in row:
                cell.border = thin_border
